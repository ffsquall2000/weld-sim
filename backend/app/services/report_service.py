"""Report generation service -- PDF, Excel, and JSON reports from Run data.

Migrated from ``ultrasonic_weld_master/plugins/reporter/`` to the new
FastAPI backend.  The service queries Run, Metric, Artifact, Project,
and SimulationCase data via SQLAlchemy async sessions and delegates
rendering to *reportlab* (PDF), *openpyxl* (Excel), and stdlib *json*
(JSON).
"""

from __future__ import annotations

import json
import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from backend.app.config import settings
from backend.app.models.artifact import Artifact
from backend.app.models.metric import Metric
from backend.app.models.project import Project
from backend.app.models.run import Run
from backend.app.models.simulation_case import SimulationCase

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

#: The 12 standard metrics tracked by the platform.
STANDARD_METRICS: List[Dict[str, Any]] = [
    {
        "name": "frequency_deviation_pct",
        "display": "Frequency Deviation",
        "unit": "%",
        "threshold_min": None,
        "threshold_max": 5.0,
    },
    {
        "name": "amplitude_uniformity",
        "display": "Amplitude Uniformity",
        "unit": "%",
        "threshold_min": 80.0,
        "threshold_max": None,
    },
    {
        "name": "stress_safety_factor",
        "display": "Stress Safety Factor",
        "unit": "",
        "threshold_min": 1.5,
        "threshold_max": None,
    },
    {
        "name": "max_von_mises_stress",
        "display": "Max Von Mises Stress",
        "unit": "MPa",
        "threshold_min": None,
        "threshold_max": 500.0,
    },
    {
        "name": "modal_frequency_hz",
        "display": "Modal Frequency",
        "unit": "Hz",
        "threshold_min": 19000.0,
        "threshold_max": 21000.0,
    },
    {
        "name": "gain_ratio",
        "display": "Gain Ratio",
        "unit": "",
        "threshold_min": 1.0,
        "threshold_max": None,
    },
    {
        "name": "contact_pressure_uniformity",
        "display": "Contact Pressure Uniformity",
        "unit": "%",
        "threshold_min": 70.0,
        "threshold_max": None,
    },
    {
        "name": "thermal_rise_c",
        "display": "Thermal Rise",
        "unit": "\u00b0C",
        "threshold_min": None,
        "threshold_max": 150.0,
    },
    {
        "name": "fatigue_cycles",
        "display": "Fatigue Cycles",
        "unit": "cycles",
        "threshold_min": 1_000_000,
        "threshold_max": None,
    },
    {
        "name": "energy_efficiency",
        "display": "Energy Efficiency",
        "unit": "%",
        "threshold_min": 60.0,
        "threshold_max": None,
    },
    {
        "name": "weld_strength_estimate",
        "display": "Weld Strength Estimate",
        "unit": "N",
        "threshold_min": 500.0,
        "threshold_max": None,
    },
    {
        "name": "coupling_loss_db",
        "display": "Coupling Loss",
        "unit": "dB",
        "threshold_min": None,
        "threshold_max": 3.0,
    },
]

_METRIC_LOOKUP: Dict[str, Dict[str, Any]] = {m["name"]: m for m in STANDARD_METRICS}


def _reports_dir() -> Path:
    """Return (and create) the reports output directory."""
    base = Path(settings.STORAGE_PATH) / "reports"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _evaluate_status(
    value: float, threshold_min: Optional[float], threshold_max: Optional[float]
) -> str:
    """Return 'pass', 'warn', or 'fail' based on threshold evaluation."""
    if threshold_min is not None and value < threshold_min:
        # Slightly below -> warn, far below -> fail
        if threshold_min != 0 and value < threshold_min * 0.8:
            return "fail"
        return "warn"
    if threshold_max is not None and value > threshold_max:
        if threshold_max != 0 and value > threshold_max * 1.2:
            return "fail"
        return "warn"
    return "pass"


# ---------------------------------------------------------------------------
# Data fetching helpers
# ---------------------------------------------------------------------------


async def _fetch_run_with_relations(
    session: AsyncSession, run_id: uuid.UUID
) -> Run:
    """Load a Run together with its metrics, artifacts, simulation_case, and project."""
    stmt = (
        select(Run)
        .where(Run.id == run_id)
        .options(
            selectinload(Run.metrics),
            selectinload(Run.artifacts),
            selectinload(Run.simulation_case).selectinload(SimulationCase.project),
        )
    )
    result = await session.execute(stmt)
    run = result.scalar_one_or_none()
    if run is None:
        raise ValueError(f"Run {run_id} not found")
    return run


def _build_metric_rows(metrics: Sequence[Metric]) -> List[Dict[str, Any]]:
    """Convert DB Metric objects into enriched row dicts aligned with the 12 standard metrics."""
    metric_map: Dict[str, Metric] = {m.metric_name: m for m in metrics}
    rows: List[Dict[str, Any]] = []
    for spec in STANDARD_METRICS:
        m = metric_map.get(spec["name"])
        if m is not None:
            status = _evaluate_status(m.value, spec["threshold_min"], spec["threshold_max"])
            rows.append(
                {
                    "metric_name": spec["name"],
                    "display_name": spec["display"],
                    "value": m.value,
                    "unit": m.unit or spec["unit"],
                    "status": status,
                    "threshold_min": spec["threshold_min"],
                    "threshold_max": spec["threshold_max"],
                }
            )
        else:
            rows.append(
                {
                    "metric_name": spec["name"],
                    "display_name": spec["display"],
                    "value": None,
                    "unit": spec["unit"],
                    "status": "n/a",
                    "threshold_min": spec["threshold_min"],
                    "threshold_max": spec["threshold_max"],
                }
            )
    return rows


def _screenshot_paths_from_artifacts(artifacts: Sequence[Artifact]) -> List[str]:
    """Extract image file paths from artifacts that look like screenshots."""
    paths: List[str] = []
    for a in artifacts:
        if a.artifact_type in ("screenshot", "vtk_screenshot", "image"):
            paths.append(a.file_path)
        elif a.mime_type and a.mime_type.startswith("image/"):
            paths.append(a.file_path)
    return paths


# ---------------------------------------------------------------------------
# Structured run data (shared across generators)
# ---------------------------------------------------------------------------


def _run_to_report_data(run: Run) -> Dict[str, Any]:
    """Convert a fully-loaded Run ORM object to a plain dict for report rendering."""
    sim_case: SimulationCase = run.simulation_case
    project: Project = sim_case.project

    return {
        "run_id": str(run.id),
        "status": run.status,
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "completed_at": run.completed_at.isoformat() if run.completed_at else None,
        "compute_time_s": run.compute_time_s,
        "input_snapshot": run.input_snapshot or {},
        "project_name": project.name,
        "project_description": project.description,
        "application_type": project.application_type,
        "simulation_name": sim_case.name,
        "simulation_description": sim_case.description,
        "analysis_type": sim_case.analysis_type,
        "solver_backend": sim_case.solver_backend,
        "configuration": sim_case.configuration or {},
        "boundary_conditions": sim_case.boundary_conditions or {},
        "material_assignments": sim_case.material_assignments or {},
        "metric_rows": _build_metric_rows(run.metrics),
        "screenshot_paths": _screenshot_paths_from_artifacts(run.artifacts),
    }


# ====================================================================
# PDF Generator
# ====================================================================


def _generate_pdf(
    runs_data: List[Dict[str, Any]],
    output_dir: Path,
    title: Optional[str],
    include_screenshots: bool,
) -> str:
    """Generate a PDF report and return the file path."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable,
        Image,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"report_{timestamp}.pdf"
    path = str(output_dir / filename)

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=20, spaceAfter=14)
    heading_style = ParagraphStyle("SectionHead", parent=styles["Heading2"], fontSize=14, spaceAfter=8)
    subheading_style = ParagraphStyle("SubHead", parent=styles["Heading3"], fontSize=12, spaceAfter=6)
    normal = styles["Normal"]

    elements: list[Any] = []
    is_comparison = len(runs_data) > 1
    primary = runs_data[0]

    # ---- Helpers ----
    def _make_table(data: list, header: bool = False, col_widths: Any = None) -> Table:
        t = Table(data, colWidths=col_widths)
        cmds: list[Any] = [
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
        if header:
            cmds.extend([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ])
        t.setStyle(TableStyle(cmds))
        return t

    _STATUS_COLORS = {
        "pass": colors.HexColor("#27AE60"),
        "warn": colors.HexColor("#F39C12"),
        "fail": colors.HexColor("#E74C3C"),
        "n/a": colors.grey,
    }

    # ---- Title page ----
    report_title = title or f"Simulation Report - {primary['project_name']}"
    elements.append(Spacer(1, 30 * mm))
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 6 * mm))

    info_data = [
        ["Project", primary["project_name"]],
        ["Application Type", primary["application_type"]],
        ["Simulation Case", primary["simulation_name"]],
        ["Analysis Type", primary["analysis_type"]],
        ["Solver Backend", primary["solver_backend"]],
        ["Generated", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")],
    ]
    if is_comparison:
        info_data.append(["Runs Compared", str(len(runs_data))])
    else:
        info_data.append(["Run ID", primary["run_id"]])
        info_data.append(["Run Status", primary["status"]])
        if primary["compute_time_s"] is not None:
            info_data.append(["Compute Time", f"{primary['compute_time_s']:.1f} s"])

    elements.append(_make_table(info_data, col_widths=[55 * mm, 110 * mm]))
    elements.append(PageBreak())

    # ---- Per-run sections (or single run) ----
    for idx, rd in enumerate(runs_data):
        run_label = f"Run {idx + 1}: {rd['run_id'][:8]}..." if is_comparison else ""

        # Section: Parameters
        if run_label:
            elements.append(Paragraph(run_label, heading_style))
            elements.append(Spacer(1, 3 * mm))

        elements.append(Paragraph("Parameters", heading_style if not run_label else subheading_style))

        # Solver configuration
        config = rd["configuration"]
        if config:
            config_rows = [["Setting", "Value"]]
            for k, v in config.items():
                config_rows.append([k.replace("_", " ").title(), str(v)])
            elements.append(Paragraph("Solver Configuration", subheading_style))
            elements.append(_make_table(config_rows, header=True, col_widths=[60 * mm, 100 * mm]))
            elements.append(Spacer(1, 3 * mm))

        # Materials
        materials = rd["material_assignments"]
        if materials:
            mat_rows = [["Component", "Material"]]
            for k, v in materials.items():
                mat_rows.append([k, str(v)])
            elements.append(Paragraph("Material Assignments", subheading_style))
            elements.append(_make_table(mat_rows, header=True, col_widths=[60 * mm, 100 * mm]))
            elements.append(Spacer(1, 3 * mm))

        # Boundary conditions
        bcs = rd["boundary_conditions"]
        if bcs:
            bc_rows = [["Condition", "Value"]]
            for k, v in bcs.items():
                bc_rows.append([k.replace("_", " ").title(), str(v)])
            elements.append(Paragraph("Boundary Conditions", subheading_style))
            elements.append(_make_table(bc_rows, header=True, col_widths=[60 * mm, 100 * mm]))
            elements.append(Spacer(1, 3 * mm))

        # Section: Results (metrics)
        elements.append(Paragraph("Results", heading_style if not run_label else subheading_style))
        metric_header = ["Metric", "Value", "Unit", "Status", "Threshold"]
        metric_table_data = [metric_header]
        for mr in rd["metric_rows"]:
            threshold_str = ""
            if mr["threshold_min"] is not None and mr["threshold_max"] is not None:
                threshold_str = f"{mr['threshold_min']} - {mr['threshold_max']}"
            elif mr["threshold_min"] is not None:
                threshold_str = f">= {mr['threshold_min']}"
            elif mr["threshold_max"] is not None:
                threshold_str = f"<= {mr['threshold_max']}"
            val_str = f"{mr['value']:.4g}" if mr["value"] is not None else "N/A"
            metric_table_data.append([
                mr["display_name"],
                val_str,
                mr["unit"] or "",
                mr["status"].upper(),
                threshold_str,
            ])

        t = _make_table(
            metric_table_data,
            header=True,
            col_widths=[45 * mm, 25 * mm, 20 * mm, 20 * mm, 40 * mm],
        )
        # Color-code status cells
        for row_idx, mr in enumerate(rd["metric_rows"], start=1):
            color = _STATUS_COLORS.get(mr["status"], colors.grey)
            t.setStyle(TableStyle([
                ("TEXTCOLOR", (3, row_idx), (3, row_idx), color),
                ("FONTNAME", (3, row_idx), (3, row_idx), "Helvetica-Bold"),
            ]))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

        # Section: Risk Assessment
        fail_metrics = [mr for mr in rd["metric_rows"] if mr["status"] == "fail"]
        warn_metrics = [mr for mr in rd["metric_rows"] if mr["status"] == "warn"]
        elements.append(Paragraph("Risk Assessment", heading_style if not run_label else subheading_style))

        risk_data = [["Category", "Level"]]
        if fail_metrics:
            risk_data.append(["Overall Risk", "HIGH"])
        elif warn_metrics:
            risk_data.append(["Overall Risk", "MEDIUM"])
        else:
            risk_data.append(["Overall Risk", "LOW"])

        for mr in fail_metrics:
            risk_data.append([mr["display_name"], "FAIL"])
        for mr in warn_metrics:
            risk_data.append([mr["display_name"], "WARN"])

        risk_table = _make_table(risk_data, header=True, col_widths=[80 * mm, 80 * mm])
        # Color-code risk rows
        for row_idx in range(1, len(risk_data)):
            level = risk_data[row_idx][1]
            if level in ("HIGH", "FAIL"):
                color = colors.HexColor("#E74C3C")
            elif level in ("MEDIUM", "WARN"):
                color = colors.HexColor("#F39C12")
            else:
                color = colors.HexColor("#27AE60")
            risk_table.setStyle(TableStyle([
                ("TEXTCOLOR", (1, row_idx), (1, row_idx), color),
                ("FONTNAME", (1, row_idx), (1, row_idx), "Helvetica-Bold"),
            ]))
        elements.append(risk_table)
        elements.append(Spacer(1, 4 * mm))

        # Section: Recommendations
        elements.append(Paragraph("Recommendations", heading_style if not run_label else subheading_style))
        recommendations = _generate_recommendations(rd["metric_rows"])
        if recommendations:
            for rec in recommendations:
                elements.append(Paragraph(f"\u2022 {rec}", normal))
        else:
            elements.append(Paragraph("All metrics within acceptable range. No action required.", normal))
        elements.append(Spacer(1, 4 * mm))

        # Screenshots placeholder
        if include_screenshots and rd["screenshot_paths"]:
            elements.append(Paragraph("Screenshots", heading_style if not run_label else subheading_style))
            for sp in rd["screenshot_paths"]:
                if os.path.isfile(sp):
                    try:
                        img = Image(sp, width=150 * mm, height=100 * mm)
                        elements.append(img)
                        elements.append(Spacer(1, 3 * mm))
                    except Exception:
                        elements.append(Paragraph(f"[Image could not be loaded: {sp}]", normal))
                else:
                    elements.append(Paragraph(f"[Image not found: {sp}]", normal))
            elements.append(Spacer(1, 4 * mm))

        if is_comparison and idx < len(runs_data) - 1:
            elements.append(PageBreak())

    # ---- Comparison table (multi-run) ----
    if is_comparison:
        elements.append(PageBreak())
        elements.append(Paragraph("Comparison Table", heading_style))
        comp_header = ["Metric", "Unit"] + [
            f"Run {i + 1}" for i in range(len(runs_data))
        ]
        comp_data = [comp_header]
        for spec in STANDARD_METRICS:
            row = [spec["display"], spec["unit"]]
            for rd in runs_data:
                mr = next(
                    (m for m in rd["metric_rows"] if m["metric_name"] == spec["name"]),
                    None,
                )
                if mr and mr["value"] is not None:
                    row.append(f"{mr['value']:.4g}")
                else:
                    row.append("N/A")
            comp_data.append(row)
        col_w = [45 * mm, 15 * mm] + [
            max(20, 120 // len(runs_data)) * mm for _ in runs_data
        ]
        elements.append(_make_table(comp_data, header=True, col_widths=col_w))

    # ---- Footer ----
    elements.append(Spacer(1, 8 * mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(
        Paragraph(
            "Generated by WeldSim v2.0.0",
            ParagraphStyle("Footer", parent=normal, fontSize=8, textColor=colors.grey),
        )
    )

    doc.build(elements)
    return path


# ====================================================================
# Excel Generator
# ====================================================================


def _generate_excel(
    runs_data: List[Dict[str, Any]],
    output_dir: Path,
    title: Optional[str],
) -> str:
    """Generate an Excel report and return the file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14)
    BOLD_FONT = Font(bold=True, size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    STATUS_FILLS = {
        "pass": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "warn": PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        "fail": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        "n/a": PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid"),
    }
    STATUS_FONTS = {
        "pass": Font(color="27AE60", bold=True),
        "warn": Font(color="F39C12", bold=True),
        "fail": Font(color="E74C3C", bold=True),
        "n/a": Font(color="808080"),
    }

    wb = Workbook()
    is_comparison = len(runs_data) > 1
    primary = runs_data[0]

    def _write_header_row(ws: Any, row: int, headers: List[str]) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    def _auto_width(ws: Any, min_width: int = 12) -> None:
        for col in ws.columns:
            max_len = min_width
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[col_letter].width = min(max_len, 50)

    # ---- Sheet 1: Summary ----
    ws_sum = wb.active
    ws_sum.title = "Summary"
    report_title = title or f"Simulation Report - {primary['project_name']}"
    ws_sum["A1"] = report_title
    ws_sum["A1"].font = TITLE_FONT
    ws_sum.merge_cells("A1:D1")

    summary_rows = [
        ("Project", primary["project_name"]),
        ("Application Type", primary["application_type"]),
        ("Simulation Case", primary["simulation_name"]),
        ("Analysis Type", primary["analysis_type"]),
        ("Solver Backend", primary["solver_backend"]),
        ("Generated", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    if not is_comparison:
        summary_rows.extend([
            ("Run ID", primary["run_id"]),
            ("Run Status", primary["status"]),
            ("Compute Time (s)", str(primary["compute_time_s"]) if primary["compute_time_s"] else "N/A"),
        ])
    else:
        summary_rows.append(("Runs Compared", str(len(runs_data))))

    row = 3
    ws_sum.cell(row=row, column=1, value="Field").font = BOLD_FONT
    ws_sum.cell(row=row, column=2, value="Value").font = BOLD_FONT
    row += 1
    for label, val in summary_rows:
        ws_sum.cell(row=row, column=1, value=label)
        ws_sum.cell(row=row, column=2, value=val)
        for c in range(1, 3):
            ws_sum.cell(row=row, column=c).border = THIN_BORDER
        row += 1
    _auto_width(ws_sum)

    # ---- Sheet 2: Parameters ----
    ws_params = wb.create_sheet("Parameters")
    ws_params["A1"] = "Input Parameters"
    ws_params["A1"].font = TITLE_FONT

    param_row = 3
    _write_header_row(ws_params, param_row, ["Category", "Parameter", "Value"])
    param_row += 1

    for rd in runs_data:
        run_label = rd["run_id"][:8] if is_comparison else ""
        # Solver configuration
        for k, v in rd["configuration"].items():
            ws_params.cell(row=param_row, column=1, value=f"Config {run_label}".strip())
            ws_params.cell(row=param_row, column=2, value=k.replace("_", " ").title())
            ws_params.cell(row=param_row, column=3, value=str(v))
            for c in range(1, 4):
                ws_params.cell(row=param_row, column=c).border = THIN_BORDER
            param_row += 1
        # Material assignments
        for k, v in rd["material_assignments"].items():
            ws_params.cell(row=param_row, column=1, value=f"Material {run_label}".strip())
            ws_params.cell(row=param_row, column=2, value=k)
            ws_params.cell(row=param_row, column=3, value=str(v))
            for c in range(1, 4):
                ws_params.cell(row=param_row, column=c).border = THIN_BORDER
            param_row += 1
        # Boundary conditions
        for k, v in rd["boundary_conditions"].items():
            ws_params.cell(row=param_row, column=1, value=f"BC {run_label}".strip())
            ws_params.cell(row=param_row, column=2, value=k.replace("_", " ").title())
            ws_params.cell(row=param_row, column=3, value=str(v))
            for c in range(1, 4):
                ws_params.cell(row=param_row, column=c).border = THIN_BORDER
            param_row += 1
    _auto_width(ws_params)

    # ---- Sheet 3: Metrics ----
    ws_met = wb.create_sheet("Metrics")
    ws_met["A1"] = "Simulation Metrics"
    ws_met["A1"].font = TITLE_FONT

    met_row = 3
    _write_header_row(ws_met, met_row, [
        "Metric", "Value", "Unit", "Status", "Min Threshold", "Max Threshold",
    ])
    met_row += 1

    for rd in runs_data:
        if is_comparison:
            ws_met.cell(row=met_row, column=1, value=f"--- Run {rd['run_id'][:8]} ---").font = BOLD_FONT
            met_row += 1
        for mr in rd["metric_rows"]:
            ws_met.cell(row=met_row, column=1, value=mr["display_name"])
            val_cell = ws_met.cell(
                row=met_row, column=2,
                value=mr["value"] if mr["value"] is not None else "N/A",
            )
            ws_met.cell(row=met_row, column=3, value=mr["unit"] or "")
            status_cell = ws_met.cell(row=met_row, column=4, value=mr["status"].upper())
            status_cell.fill = STATUS_FILLS.get(mr["status"], STATUS_FILLS["n/a"])
            status_cell.font = STATUS_FONTS.get(mr["status"], STATUS_FONTS["n/a"])
            ws_met.cell(
                row=met_row, column=5,
                value=mr["threshold_min"] if mr["threshold_min"] is not None else "",
            )
            ws_met.cell(
                row=met_row, column=6,
                value=mr["threshold_max"] if mr["threshold_max"] is not None else "",
            )
            for c in range(1, 7):
                ws_met.cell(row=met_row, column=c).border = THIN_BORDER
            met_row += 1
    _auto_width(ws_met)

    # ---- Sheet 4: Comparison (only for multi-run) ----
    if is_comparison:
        ws_comp = wb.create_sheet("Comparison")
        ws_comp["A1"] = "Multi-Run Comparison"
        ws_comp["A1"].font = TITLE_FONT

        comp_row = 3
        comp_headers = ["Metric", "Unit"] + [f"Run {i + 1} ({rd['run_id'][:8]})" for i, rd in enumerate(runs_data)]
        _write_header_row(ws_comp, comp_row, comp_headers)
        comp_row += 1

        for spec in STANDARD_METRICS:
            ws_comp.cell(row=comp_row, column=1, value=spec["display"])
            ws_comp.cell(row=comp_row, column=2, value=spec["unit"])
            for ri, rd in enumerate(runs_data):
                mr = next(
                    (m for m in rd["metric_rows"] if m["metric_name"] == spec["name"]),
                    None,
                )
                cell = ws_comp.cell(
                    row=comp_row,
                    column=3 + ri,
                    value=mr["value"] if mr and mr["value"] is not None else "N/A",
                )
                if mr and mr["status"] in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[mr["status"]]
                    cell.font = STATUS_FONTS.get(mr["status"], Font())
                cell.border = THIN_BORDER
            for c in range(1, 3):
                ws_comp.cell(row=comp_row, column=c).border = THIN_BORDER
            comp_row += 1
        _auto_width(ws_comp)

    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"report_{timestamp}.xlsx"
    path = str(output_dir / filename)
    wb.save(path)
    return path


# ====================================================================
# JSON Generator
# ====================================================================


def _generate_json(
    runs_data: List[Dict[str, Any]],
    output_dir: Path,
    title: Optional[str],
) -> str:
    """Generate a JSON export and return the file path."""
    report: Dict[str, Any] = {
        "report_format": "json",
        "title": title or "Simulation Report",
        "generated_at": datetime.now(tz=timezone.utc).isoformat(),
        "runs": runs_data,
    }
    if len(runs_data) > 1:
        # Add a flat comparison section
        comparison: Dict[str, Any] = {}
        for spec in STANDARD_METRICS:
            comparison[spec["name"]] = {}
            for i, rd in enumerate(runs_data):
                mr = next(
                    (m for m in rd["metric_rows"] if m["metric_name"] == spec["name"]),
                    None,
                )
                comparison[spec["name"]][rd["run_id"]] = (
                    mr["value"] if mr and mr["value"] is not None else None
                )
        report["comparison"] = comparison

    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"report_{timestamp}.json"
    path = str(output_dir / filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)
    return path


# ====================================================================
# Recommendations engine
# ====================================================================


def _generate_recommendations(metric_rows: List[Dict[str, Any]]) -> List[str]:
    """Produce textual recommendations based on metric status."""
    recs: List[str] = []
    for mr in metric_rows:
        if mr["status"] == "fail":
            recs.append(
                f"CRITICAL: {mr['display_name']} is out of acceptable range "
                f"(value={mr['value']}, threshold "
                f"min={mr['threshold_min']}, max={mr['threshold_max']}). "
                "Immediate review and parameter adjustment required."
            )
        elif mr["status"] == "warn":
            recs.append(
                f"WARNING: {mr['display_name']} is approaching threshold limits "
                f"(value={mr['value']}). Consider adjusting parameters."
            )
    return recs


# ====================================================================
# Public API
# ====================================================================


async def generate_report(
    session: AsyncSession,
    run_ids: List[str],
    fmt: str,
    include_screenshots: bool = False,
    title: Optional[str] = None,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """Generate one or more reports and return metadata.

    Parameters
    ----------
    session:
        Async SQLAlchemy session.
    run_ids:
        List of Run UUID strings.
    fmt:
        ``"pdf"``, ``"excel"``, ``"json"``, or ``"all"``.
    include_screenshots:
        If *True*, attempt to embed screenshots from artifacts.
    title:
        Optional custom report title.
    output_dir:
        Optional override for the report output directory.

    Returns
    -------
    dict
        ``{"status": "ok", "format": ..., "file_path": ..., "file_size_bytes": ...,
           "generated_at": ...}``
        or for ``"all"`` format a ``files`` dict with each format.
    """
    out = Path(output_dir) if output_dir else _reports_dir()
    out.mkdir(parents=True, exist_ok=True)

    # Fetch all runs and validate status (BUG-9 fix)
    runs_data: List[Dict[str, Any]] = []
    for rid in run_ids:
        run = await _fetch_run_with_relations(session, uuid.UUID(rid))
        if run.status not in ("completed", "failed"):
            raise ValueError(
                f"Cannot generate report for run {rid} with status '{run.status}'. "
                f"Run must be 'completed' or 'failed'."
            )
        runs_data.append(_run_to_report_data(run))

    now = datetime.now(tz=timezone.utc)

    if fmt == "all":
        files: Dict[str, str] = {}
        sizes: Dict[str, int] = {}
        for f, gen in [
            ("pdf", lambda: _generate_pdf(runs_data, out, title, include_screenshots)),
            ("excel", lambda: _generate_excel(runs_data, out, title)),
            ("json", lambda: _generate_json(runs_data, out, title)),
        ]:
            path = gen()
            files[f] = path
            sizes[f] = os.path.getsize(path)
        return {
            "status": "ok",
            "files": files,
            "file_sizes": sizes,
            "generated_at": now,
        }

    if fmt == "pdf":
        path = _generate_pdf(runs_data, out, title, include_screenshots)
    elif fmt == "excel":
        path = _generate_excel(runs_data, out, title)
    elif fmt == "json":
        path = _generate_json(runs_data, out, title)
    else:
        raise ValueError(f"Unsupported report format: {fmt}")

    return {
        "status": "ok",
        "format": fmt,
        "file_path": path,
        "file_size_bytes": os.path.getsize(path),
        "generated_at": now,
    }


def list_reports(output_dir: Optional[str] = None) -> List[Dict[str, Any]]:
    """List all report files in the output directory."""
    out = Path(output_dir) if output_dir else _reports_dir()
    if not out.is_dir():
        return []

    items: List[Dict[str, Any]] = []
    for fp in sorted(out.iterdir()):
        if fp.is_file() and fp.suffix in (".pdf", ".xlsx", ".json"):
            fmt_map = {".pdf": "pdf", ".xlsx": "excel", ".json": "json"}
            stat = fp.stat()
            items.append({
                "filename": fp.name,
                "format": fmt_map.get(fp.suffix, "unknown"),
                "file_size_bytes": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_ctime, tz=timezone.utc),
            })
    return items


def get_report_path(filename: str, output_dir: Optional[str] = None) -> Optional[str]:
    """Return the full path of a report file if it exists."""
    out = Path(output_dir) if output_dir else _reports_dir()
    fp = out / filename
    if fp.is_file():
        return str(fp)
    return None


def delete_report(filename: str, output_dir: Optional[str] = None) -> bool:
    """Delete a report file. Returns True if deleted, False if not found."""
    out = Path(output_dir) if output_dir else _reports_dir()
    fp = out / filename
    if fp.is_file():
        fp.unlink()
        return True
    return False
