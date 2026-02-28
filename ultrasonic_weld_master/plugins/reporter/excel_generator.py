"""Excel report generator using openpyxl."""
from __future__ import annotations

import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ultrasonic_weld_master.core.models import WeldRecipe, ValidationResult


_HEADER_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class ExcelGenerator:
    def export(self, recipe: WeldRecipe, validation: Optional[ValidationResult] = None,
               output_dir: str = ".") -> str:
        wb = Workbook()

        self._write_summary_sheet(wb.active, recipe, validation)
        self._write_parameters_sheet(wb.create_sheet("Parameters"), recipe)
        self._write_safety_sheet(wb.create_sheet("Safety"), recipe)
        if validation:
            self._write_validation_sheet(wb.create_sheet("Validation"), validation)

        filename = "report_%s_%s.xlsx" % (recipe.recipe_id, datetime.now().strftime("%Y%m%d_%H%M%S"))
        path = os.path.join(output_dir, filename)
        wb.save(path)
        return path

    def _write_summary_sheet(self, ws, recipe, validation):
        ws.title = "Summary"
        ws["A1"] = "Ultrasonic Welding Parameter Report"
        ws["A1"].font = _TITLE_FONT
        ws.merge_cells("A1:D1")

        ws["A3"] = "Recipe ID"
        ws["B3"] = recipe.recipe_id
        ws["A4"] = "Application"
        ws["B4"] = recipe.application
        ws["A5"] = "Generated"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = 7
        ws.cell(row=row, column=1, value="Input Parameter").font = _HEADER_FONT
        ws.cell(row=row, column=2, value="Value").font = _HEADER_FONT
        row += 1
        for k, v in recipe.inputs.items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v))
            row += 1

        if validation:
            row += 1
            ws.cell(row=row, column=1, value="Validation Status").font = _HEADER_FONT
            ws.cell(row=row, column=2, value=validation.status.value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _write_parameters_sheet(self, ws, recipe):
        headers = ["Parameter", "Value", "Min (Safe)", "Max (Safe)", "Unit"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        units = {
            "amplitude_um": "um", "pressure_n": "N", "pressure_mpa": "MPa",
            "energy_j": "J", "time_ms": "ms", "frequency_khz": "kHz",
        }

        row = 2
        for param, value in recipe.parameters.items():
            ws.cell(row=row, column=1, value=param)
            # Convert non-primitive types (dict, list) to string for Excel compatibility
            cell_value = str(value) if isinstance(value, (dict, list, set)) else value
            ws.cell(row=row, column=2, value=cell_value)
            sw = recipe.safety_window.get(param)
            if sw and isinstance(sw, (list, tuple)) and len(sw) >= 2:
                ws.cell(row=row, column=3, value=sw[0])
                ws.cell(row=row, column=4, value=sw[1])
            ws.cell(row=row, column=5, value=units.get(param, ""))
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = _THIN_BORDER
            row += 1

        for col_letter in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = 18

    def _write_safety_sheet(self, ws, recipe):
        ws["A1"] = "Risk Assessment"
        ws["A1"].font = _TITLE_FONT

        row = 3
        for k, v in recipe.risk_assessment.items():
            ws.cell(row=row, column=1, value=k)
            cell = ws.cell(row=row, column=2, value=v)
            if v == "high":
                cell.font = Font(color="FF0000", bold=True)
            elif v == "medium":
                cell.font = Font(color="FF8C00", bold=True)
            else:
                cell.font = Font(color="008000")
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Recommendations").font = _HEADER_FONT
        row += 1
        for rec in recipe.recommendations:
            ws.cell(row=row, column=1, value=rec)
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _write_validation_sheet(self, ws, validation):
        ws["A1"] = "Validation Results"
        ws["A1"].font = _TITLE_FONT
        ws["A3"] = "Overall Status"
        ws["B3"] = validation.status.value

        row = 5
        for name, result in validation.validators.items():
            ws.cell(row=row, column=1, value=name).font = _HEADER_FONT
            ws.cell(row=row, column=2, value=result.get("status", ""))
            row += 1
            for msg in result.get("messages", []):
                ws.cell(row=row, column=2, value=msg)
                row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60
