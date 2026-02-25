from __future__ import annotations
import os
import pytest
from openpyxl import load_workbook
from ultrasonic_weld_master.core.models import WeldRecipe, ValidationResult, ValidationStatus
from ultrasonic_weld_master.plugins.reporter.excel_generator import ExcelGenerator

def _make_recipe():
    return WeldRecipe(
        recipe_id="EX001", application="li_battery_tab",
        inputs={"upper_material": "Al", "weld_area_mm2": 125.0},
        parameters={"amplitude_um": 30.0, "pressure_n": 37.5, "pressure_mpa": 0.3,
                     "energy_j": 60.0, "time_ms": 200, "frequency_khz": 20.0},
        safety_window={"amplitude_um": [25.5, 34.5], "energy_j": [48.0, 78.0]},
        risk_assessment={"overweld_risk": "low", "underweld_risk": "low", "perforation_risk": "low"},
        recommendations=["Start at 80% energy"],
    )

class TestExcelGenerator:
    def test_export_creates_file(self, tmp_path):
        recipe = _make_recipe()
        gen = ExcelGenerator()
        path = gen.export(recipe, output_dir=str(tmp_path))
        assert os.path.exists(path)
        assert path.endswith(".xlsx")

    def test_export_has_sheets(self, tmp_path):
        recipe = _make_recipe()
        validation = ValidationResult(
            status=ValidationStatus.PASS,
            validators={"physics": {"status": "pass", "messages": []}},
        )
        gen = ExcelGenerator()
        path = gen.export(recipe, validation, str(tmp_path))
        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Parameters" in sheet_names
        assert "Safety" in sheet_names
        assert "Validation" in sheet_names

    def test_parameters_sheet_content(self, tmp_path):
        recipe = _make_recipe()
        gen = ExcelGenerator()
        path = gen.export(recipe, output_dir=str(tmp_path))
        wb = load_workbook(path)
        ws = wb["Parameters"]
        assert ws.cell(row=1, column=1).value == "Parameter"
        assert ws.cell(row=2, column=1).value == "amplitude_um"
        assert ws.cell(row=2, column=2).value == 30.0
